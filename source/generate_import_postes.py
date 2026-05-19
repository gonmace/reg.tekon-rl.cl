"""
Genera source/import_postes_RL.xlsx listo para importar en sitios via SiteResource.

Fuente: 28 POSTES RL.xlsx (raíz del proyecto)
Uso:    python3 source/generate_import_postes.py
"""

import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_FILE = os.path.join(BASE_DIR, "28 POSTES RL.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "source", "import_postes_RL.xlsx")

HEADERS = [
    "PTI ID",
    "Operador ID",
    "Nombre del Sitio",
    "Tipo de sitio",
    "Latitud Mandato",
    "Longitud Mandato",
    "Latitud Ingeniería",
    "Longitud Ingeniería",
    "Latitud Construcción",
    "Longitud Construcción",
    "Altura (m)",
    "Region / Provincia",
    "Comuna / Municipio",
    "Empresa de Energía",
]


def main():
    wb_src = openpyxl.load_workbook(SOURCE_FILE)
    ws_src = wb_src.active

    src_headers = [cell.value for cell in ws_src[1]]
    col = {h: i for i, h in enumerate(src_headers)}

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(HEADERS)

    count = 0
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        ws_out.append([
            None,                       # PTI ID — vacío
            row[col["ID"]],             # Operador ID
            row[col["Site Name"]],      # Nombre del Sitio
            "POSTE",                    # Tipo de sitio
            row[col["LAT"]],            # Latitud Mandato
            row[col["LONG"]],           # Longitud Mandato
            None,                       # Latitud Ingeniería
            None,                       # Longitud Ingeniería
            None,                       # Latitud Construcción
            None,                       # Longitud Construcción
            None,                       # Altura (m)
            row[col["Region"]],         # Region / Provincia
            row[col["Comuna"]],         # Comuna / Municipio
            row[col["EMPRESA FINAL"]],  # Empresa de Energía
        ])
        count += 1

    wb_out.save(OUTPUT_FILE)
    print(f"Generado: {OUTPUT_FILE}  ({count} filas)")


if __name__ == "__main__":
    main()
